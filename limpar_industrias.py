import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ============================================
# SCRIPT PARA LIMPAR A ABA "INDUSTRIAS"
# ============================================

# Ajuste este caminho se necessário
caminho_planilha = r"C:\Users\sgcar\OneDrive\Desktop\APP DESTRO\1 - Python\Programa_Destro-04-03.xlsx"

def limpar_industria(val):
    """Extrai apenas o nome entre as primeiras aspas duplas"""
    val_str = str(val)
    match = re.search(r'\"(.*?)\"', val_str)
    if match:
        return match.group(1).strip()
    return val_str.replace("(", "").replace(")", "").replace("'", "").replace('"', "").split(",")[0].strip()

try:
    print("=" * 60)
    print("🚀 INICIANDO LIMPEZA DA ABA INDUSTRIAS")
    print("=" * 60)
    
    # Lê a aba Industrias
    df = pd.read_excel(caminho_planilha, sheet_name="Industrias", header=None)
    
    print(f"\n✅ Aba carregada com sucesso!")
    print(f"📊 Total de linhas: {len(df)}")
    
    print(f"\n🔍 ANTES (primeiras 3 linhas):")
    for idx, val in enumerate(df[0].head(3), 1):
        print(f"   {idx}. {val}")
    
    # Limpa a coluna
    df['Nome Limpo'] = df[0].apply(limpar_industria)
    
    print(f"\n✨ DEPOIS (primeiras 3 linhas limpas):")
    for idx, val in enumerate(df['Nome Limpo'].head(3), 1):
        print(f"   {idx}. {val}")
    
    # Remove linhas vazias
    df = df[df['Nome Limpo'].str.strip() != '']
    
    print(f"\n📉 Após remover vazios: {len(df)} linhas")
    
    # Salva no Excel
    with pd.ExcelWriter(caminho_planilha, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name='Industrias', index=False, header=['Original', 'Nome Limpo'])
    
    print(f"\n✅ Arquivo salvo com sucesso!")
    print(f"📁 {caminho_planilha}")
    
    # Formatação visual (opcional)
    wb = load_workbook(caminho_planilha)
    ws = wb['Industrias']
    
    # Cabeçalhos com estilo
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ajusta a largura das colunas
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 40
    
    wb.save(caminho_planilha)
    
    print("\n🎨 Formatação aplicada (cabeçalhos e colunas ajustadas)")
    print("\n" + "=" * 60)
    print("✨ PROCESSO CONCLUÍDO COM SUCESSO!")
    print("=" * 60)
    
except FileNotFoundError:
    print(f"\n❌ ERRO: Arquivo não encontrado!")
    print(f"   Caminho esperado: {caminho_planilha}")
    print(f"\n💡 Dicas:")
    print(f"   1. Verifique se o arquivo existe neste local")
    print(f"   2. Feche o arquivo no Excel antes de rodar o script")
    print(f"   3. Verifique se o caminho está correto no script")
    
except Exception as e:
    print(f"\n❌ ERRO: {str(e)}")
    print(f"\n💡 Certifique-se que:")
    print(f"   1. O arquivo não está aberto no Excel")
    print(f"   2. Você tem permissão para editar o arquivo")
    print(f"   3. O caminho está correto")

input("\n👉 Pressione ENTER para sair...")
