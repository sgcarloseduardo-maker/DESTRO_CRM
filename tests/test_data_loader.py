from pathlib import Path

from openpyxl import Workbook

from data_loader import load_crm_data


def _create_minimal_valid_workbook(path: Path):
    wb = Workbook()

    ws_abc = wb.active
    ws_abc.title = "Curva ABC_Semanal"
    ws_abc.cell(row=1, column=2, value="TANTE: 999")
    ws_abc.cell(row=3, column=1, value="12345-0")
    ws_abc.cell(row=3, column=2, value="ABC")

    ws_ind = wb.create_sheet("Industrias")
    ws_ind.cell(row=2, column=1, value='"IND TESTE"')

    ws_marcas = wb.create_sheet("Marcas")
    ws_marcas.cell(row=2, column=1, value="MARCA TESTE")

    ws_bl = wb.create_sheet("BateuLevou")
    ws_bl.append(["DESCRICAO", "INDUSTRIA", "MARCA"])
    ws_bl.append(["PRODUTO TESTE", "IND TESTE", "MARCA TESTE"])

    ws_banco = wb.create_sheet("Banco_Dados_Semanal")
    for i in range(1, 6):
        ws_banco.cell(row=i, column=1, value="")
    ws_banco.cell(row=6, column=1, value="12345-0")
    ws_banco.cell(row=6, column=2, value="001 - CATEGORIA TESTE")
    ws_banco.cell(row=6, column=3, value="PRODUTO TESTE")
    ws_banco.cell(row=6, column=6, value=10.0)
    ws_banco.cell(row=6, column=8, value=11.0)
    ws_banco.cell(row=6, column=10, value=12.0)
    ws_banco.cell(row=6, column=12, value=13.0)
    ws_banco.cell(row=6, column=13, value="*")

    ws_tp = wb.create_sheet("TO PODENDO")
    ws_tp.cell(row=1, column=1, value="12345")

    ws_col = wb.create_sheet("COLGATE")
    ws_col.cell(row=1, column=1, value="12345")

    wb.save(path)


def test_load_crm_data_missing_file(tmp_path: Path):
    df, abc, warnings, fatal = load_crm_data(str(tmp_path / "inexistente.xlsx"))
    assert df.empty
    assert abc == []
    assert warnings == []
    assert fatal == "PLANILHA_NAO_ENCONTRADA"


def test_load_crm_data_invalid_file(tmp_path: Path):
    file_path = tmp_path / "invalid.xlsx"
    file_path.write_text("nao e excel", encoding="utf-8")
    df, abc, warnings, fatal = load_crm_data(str(file_path))
    assert df.empty
    assert abc == []
    assert warnings == []
    assert fatal == "PLANILHA_CORROMPIDA_OU_INVALIDA"


def test_load_crm_data_structure_invalid(tmp_path: Path):
    file_path = tmp_path / "estrutura_invalida.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Banco_Dados_Semanal"
    ws.cell(row=6, column=1, value="123")
    wb.save(file_path)

    df, abc, warnings, fatal = load_crm_data(str(file_path))
    assert df.empty
    assert fatal == "ESTRUTURA_BANCO_DADOS_INVALIDA"
    assert any("Abas ausentes" in w for w in warnings)


def test_load_crm_data_minimal_valid(tmp_path: Path):
    file_path = tmp_path / "valida.xlsx"
    _create_minimal_valid_workbook(file_path)

    df, abc, warnings, fatal = load_crm_data(str(file_path))
    assert fatal == ""
    assert not df.empty
    assert "Código" in df.columns
    assert "Status" not in df.columns  # status pertence ao processamento da camada app
    assert isinstance(abc, list)
    # workbook mínimo pode gerar warnings de fallback, mas não deve falhar
    assert warnings is not None
